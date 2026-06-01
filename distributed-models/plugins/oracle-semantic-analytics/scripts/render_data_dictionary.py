#!/usr/bin/env python3
"""
Generate an Excel data dictionary from the bundled semantic YAML model.

Sheets:
  Dimensions   — logical entities, fields, physical mapping, PK, restricted flag
  Measures     — measures and calculated measures with expressions and rules
  Glossary     — business terms mapped to canonical measures/dimensions
  Governance   — restricted fields, prohibited requests, privacy controls
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

try:
    import yaml
except ImportError:
    print("Missing package: PyYAML. Run /oracle-semantic-analytics:setup-analytics")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing package: openpyxl. Run /oracle-semantic-analytics:setup-analytics")
    sys.exit(1)

from analytics_config import runtime_settings

ROOT = Path(__file__).resolve().parents[1]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_model(model_path: Path) -> dict:
    return yaml.safe_load(model_path.read_text(encoding="utf-8"))


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def style_row(ws, row: int, ncols: int, alt: bool) -> None:
    fill = ALT_FILL if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = WRAP
        if fill:
            cell.fill = fill


def autofit(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 4, max_width))


def write_dimensions(wb: openpyxl.Workbook, model: dict) -> None:
    ws = wb.create_sheet("Dimensions")
    headers = [
        "Subject Area", "Entity", "Entity Label", "Logical Field",
        "Business Label", "Description", "Physical Table", "Physical Column",
        "Is PK", "Is Restricted", "Default Filter",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    subject_area = model.get("semantic_model", {}).get("name", "")
    logical = model.get("logical_model", {})
    entities = {e["id"]: e for e in logical.get("entities", [])}
    governance = model.get("governance", {})
    restricted_logical = {
        f for f in governance.get("security", {}).get("restricted_logical_fields", [])
    }

    # Build physical column lookup: logical_entity -> {logical_field -> physical_column}
    physical_lookup: dict[str, dict[str, str]] = {}
    table_lookup: dict[str, str] = {}
    for mapping in model.get("physical_mappings", []):
        for obj in mapping.get("objects", []):
            entity_id = obj.get("logical_entity", "")
            fq_name = obj.get("fully_qualified_name", "")
            table_lookup[entity_id] = fq_name
            physical_lookup[entity_id] = obj.get("columns", {})

    # Build PK lookup: logical_entity -> set of logical field names
    pk_lookup: dict[str, set] = {}
    for rel in logical.get("relationships", []):
        from_entity = rel.get("from_entity", "")
        pk_col = rel.get("from_key", "")
        if from_entity and pk_col:
            pk_lookup.setdefault(from_entity, set()).add(pk_col)
    # Also mark grain keys
    for e in logical.get("entities", []):
        for key in e.get("grain", {}).get("keys", []):
            pk_lookup.setdefault(e["id"], set()).add(key)

    # Default filters per entity
    filters = {f["id"]: f for f in logical.get("filters", [])}

    row_num = 2
    for entity in logical.get("entities", []):
        entity_id = entity["id"]
        entity_label = entity.get("label", entity.get("name", entity_id))
        cols = physical_lookup.get(entity_id, {})
        phys_table = table_lookup.get(entity_id, "")
        pks = pk_lookup.get(entity_id, set())

        for logical_field, phys_col in cols.items():
            is_pk = "Yes" if logical_field in pks else ""
            is_restricted = "Yes" if logical_field in restricted_logical else ""
            # find default filter mentioning this entity
            default_filter = ""
            for f in logical.get("filters", []):
                if f.get("entity") == entity_id:
                    default_filter = f.get("expression", {}).get("logical", "")
                    break

            ws.append([
                subject_area, entity_id, entity_label, logical_field,
                logical_field.replace("_", " ").title(), "",
                phys_table, phys_col, is_pk, is_restricted, default_filter,
            ])
            style_row(ws, row_num, len(headers), row_num % 2 == 0)
            row_num += 1

    autofit(ws)
    ws.freeze_panes = "A2"


def write_measures(wb: openpyxl.Workbook, model: dict) -> None:
    ws = wb.create_sheet("Measures")
    headers = [
        "Subject Area", "Measure Name", "Business Label", "Type",
        "Expression (Logical)", "Physical Expression",
        "Default Filters", "Allowed Dimensions",
        "Business Rules / Privacy Warning", "Format",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    subject_area = model.get("semantic_model", {}).get("name", "")
    logical = model.get("logical_model", {})

    # Physical measure expressions
    phys_measures: dict[str, str] = {}
    for mapping in model.get("physical_mappings", []):
        phys_measures.update(mapping.get("measure_expressions", {}))

    row_num = 2
    all_measures = logical.get("measures", []) + logical.get("calculated_measures", [])
    for m in all_measures:
        name = m.get("name", "")
        label = m.get("label", name)
        mtype = m.get("type", "calculated")
        expr_logical = m.get("expression", {}).get("logical", "") or m.get("formula", {}).get("logical", "")
        phys_expr = phys_measures.get(f"measure.{name}", phys_measures.get(name, ""))
        default_filters = ", ".join(m.get("default_filters", []))
        allowed_dims = ", ".join(m.get("allowed_dimensions", []))
        warnings = []
        if m.get("privacy_warning"):
            warnings.append(m["privacy_warning"])
        if m.get("reliability_warning"):
            warnings.append(m["reliability_warning"])
        rules = " | ".join(warnings)
        fmt = m.get("format", {}).get("type", "")

        ws.append([
            subject_area, name, label, mtype,
            expr_logical, phys_expr,
            default_filters, allowed_dims, rules, fmt,
        ])
        style_row(ws, row_num, len(headers), row_num % 2 == 0)
        row_num += 1

    autofit(ws)
    ws.freeze_panes = "A2"


def write_glossary(wb: openpyxl.Workbook, model: dict) -> None:
    ws = wb.create_sheet("Glossary")
    headers = [
        "Business Term", "Maps To (Canonical)", "Synonyms", "Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    row_num = 2
    for entry in model.get("business_glossary", []):
        term = entry.get("term", "")
        maps_to = entry.get("maps_to", "")
        if isinstance(maps_to, list):
            maps_to = ", ".join(maps_to)
        synonyms = ", ".join(entry.get("synonyms", []))
        notes = entry.get("notes", entry.get("description", ""))
        ws.append([term, maps_to, synonyms, notes])
        style_row(ws, row_num, len(headers), row_num % 2 == 0)
        row_num += 1

    autofit(ws)
    ws.freeze_panes = "A2"


def write_governance(wb: openpyxl.Workbook, model: dict) -> None:
    ws = wb.create_sheet("Governance")
    governance = model.get("governance", {})
    security = governance.get("security", {})
    privacy = governance.get("privacy_controls", {})

    # Section: restricted fields
    ws.append(["RESTRICTED FIELDS"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["Logical Field", "Reason"])
    style_header(ws, 2, 2)
    row_num = 3
    for field in security.get("restricted_logical_fields", []):
        ws.append([field, "PII / FERPA-sensitive"])
        style_row(ws, row_num, 2, row_num % 2 == 0)
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="PROHIBITED REQUESTS").font = Font(bold=True, size=12)
    row_num += 1
    ws.append(["Request Type"])
    style_header(ws, row_num, 1)
    row_num += 1
    for req in security.get("prohibited_requests", []):
        ws.append([req])
        style_row(ws, row_num, 1, row_num % 2 == 0)
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value="PRIVACY CONTROLS").font = Font(bold=True, size=12)
    row_num += 1
    suppression = privacy.get("small_cell_suppression", {})
    ws.append(["Setting", "Value"])
    style_header(ws, row_num, 2)
    row_num += 1
    ws.append(["Aggregate only by default", str(privacy.get("aggregate_only_by_default", ""))])
    row_num += 1
    ws.append(["Small cell suppression enabled", str(suppression.get("enabled", ""))])
    row_num += 1
    ws.append(["Suppression threshold", str(suppression.get("threshold", ""))])
    row_num += 1
    ws.append(["Suppression action", str(suppression.get("action", ""))])

    autofit(ws)


def main() -> int:
    model_path = ROOT / "assets" / "semantic_models" / "sia_term_enrollments.yaml"
    if not model_path.exists():
        print(f"Semantic model not found: {model_path}")
        return 1

    model = load_model(model_path)
    subject_area = model.get("semantic_model", {}).get("name", "oracle-semantic")

    reports_dir = Path(runtime_settings()["reports_dir"])
    reports_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{timestamp}-data-dictionary-{subject_area.lower().replace(' ', '-')}.xlsx"
    output_path = reports_dir / filename

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    write_dimensions(wb, model)
    write_measures(wb, model)
    write_glossary(wb, model)
    write_governance(wb, model)

    wb.save(output_path)
    print(output_path)

    try:
        import webbrowser
        webbrowser.open(output_path.as_uri())
    except Exception:
        pass

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
